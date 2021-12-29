# Libraaries and API
import re
import googleapiclient.discovery
from urllib.parse import parse_qs, urlparse
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side

youtube = googleapiclient.discovery.build("youtube", "v3", developerKey = "YOUR_API")


# Define time patterns for duration
hours_pattern = re.compile(r'(\d+)H')
minutes_pattern = re.compile(r'(\d+)M')
seconds_pattern = re.compile(r'(\d+)S')


# Extract playlist id from url
url = 'https://www.youtube.com/playlist?list=PLVsEAZbuBOF_nppuJ2xBGL6Ig2zG6BDv_'
query = parse_qs(urlparse(url).query, keep_blank_values=True)
playlist_id = query["list"][0]
print(f'get all playlist items links from {playlist_id}')


# Global scrip variables
workbook = load_workbook(filename="playlist.xlsx")
sheet = workbook.active
count = 1;
nextPageToken = None

posFill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
nameFill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
channelFill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
linkFill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
lenghtFill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

thinBorder = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Main code
while True:
    # First request to get IDs for all the videos
    id_request = youtube.playlistItems().list(
        part = "snippet",
        playlistId = playlist_id,
        maxResults = 50,
        pageToken = nextPageToken
    )
    id_response = id_request.execute()

    # Create array with the video IDs
    ids = []
    for item in id_response['items']:
        ids.append(item["snippet"]["resourceId"]["videoId"])

    # Second request to get duration of the video too
    vid_request = youtube.videos().list(
        part = "snippet, contentDetails",
        id = ','.join(ids)
    )
    vid_response = vid_request.execute()

    # Print information about videos
    for video in vid_response["items"]:
        duration = video["contentDetails"]["duration"]

        minutes = minutes_pattern.search(duration)
        seconds = seconds_pattern.search(duration)
        minutes = int(minutes.group(1)) if minutes else 0
        seconds = int(seconds.group(1)) if seconds else 0

        print(f'{video["snippet"]["title"]} - %02d:%02d' % (minutes, seconds))
        print('- ' + f'{video["snippet"]["channelTitle"]}')
        print('- ' + f'https://www.youtube.com/watch?v={video["id"]}&t=0s' + '\n')

    # Carry information to excel file
    for video in vid_response["items"]:
        duration = video["contentDetails"]["duration"]

        minutes = minutes_pattern.search(duration)
        seconds = seconds_pattern.search(duration)
        minutes = int(minutes.group(1)) if minutes else 0
        seconds = int(seconds.group(1)) if seconds else 0

        sheet["A" + str(count+1)] = str(count)
        sheet["A" + str(count+1)].fill = posFill
        sheet["A" + str(count+1)].border = thinBorder

        sheet["B" + str(count+1)] = video["snippet"]["title"]
        sheet["B" + str(count+1)].fill = nameFill
        sheet["B" + str(count+1)].border = thinBorder

        sheet["C" + str(count+1)] = video["snippet"]["channelTitle"]
        sheet["C" + str(count+1)].fill = channelFill
        sheet["c" + str(count+1)].border = thinBorder

        sheet["D" + str(count+1)] = f'https://www.youtube.com/watch?v={video["id"]}&t=0s'
        sheet["D" + str(count+1)].fill = linkFill
        sheet["D" + str(count+1)].border = thinBorder

        sheet["E" + str(count+1)] = "%02d:%02d" % (minutes, seconds)
        sheet["E" + str(count+1)].fill = lenghtFill
        sheet["E" + str(count+1)].border = thinBorder

        count += 1
        
    nextPageToken = id_response.get('nextPageToken')

    if not nextPageToken:
        break

workbook.save(filename="playlistDone.xlsx")


